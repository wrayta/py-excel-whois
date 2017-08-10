from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import whois
from tld import get_tld
from Tkinter import *
import tkFileDialog
import os.path

class input_window:

    def __init__(self, master):
        self.master = master
        self.frame = Frame(self.master)
        # For the 'artist roster' file
        self.artist_roster_prompt = Label(self.frame, text='Please choose the artist roster file (.xlsx): ', width=60)
        self.artist_roster_prompt.grid(row=0)
        self.artist_roster_button = Button(self.frame, text='Artist Roster', width=30,
                                      command=self.artist_roster_file_handler)
        self.artist_roster_button.grid(row=0, column=1, pady=(10, 0))

        self.artist_roster_filename_text_field = Label(self.frame, height=1, width=100)
        self.artist_roster_filename_text_field.grid(row=1, pady=(0, 10))

        # For the 'old domain' file
        self.old_domain_prompt = Label(self.frame, text='Please choose the most recent domain file (.xlsx): ', width=60)
        self.old_domain_prompt.grid(row=2)
        self.old_domain_button = Button(self.frame, text='Most Recent Domain File', width=30,
                                   command=self.old_domain_file_handler)
        self.old_domain_button.grid(row=2, column=1)

        self.old_domain_filename_text_field = Label(self.frame, height=1, width=100)
        self.old_domain_filename_text_field.grid(row=3, pady=(0, 10))

        # For the 'new domain' file directory
        self.new_domain_directory_prompt = Label(self.frame, text='Please choose where to save the new domain file: ', width=60)
        self.new_domain_directory_prompt.grid(row=4)
        self.new_domain_directory_button = Button(self.frame, text='Directory to Save New Domain File', width=30,
                                             command=self.open_directory_handler)
        self.new_domain_directory_button.grid(row=4, column=1)

        self.new_domain_directory_text_field = Label(self.frame, height=1, width=100)
        self.new_domain_directory_text_field.grid(row=5, pady=(0, 10))

        # For the 'new domain' file name
        self.new_domain_filename_prompt = Label(self.frame, text='Name the new domain file: ', width=60)
        self.new_domain_filename_prompt.grid(row=6)

        self.new_domain_filename_entry = Entry(self.frame, width=30)
        self.new_domain_filename_entry.grid(row=6, column=1)

        self.excel_label = Label(self.frame, text='.xlsx')
        self.excel_label.grid(row=6, column=2)

        # Submit the info and run the rest of the program
        self.submit_button = Button(self.frame, text='Submit', width=30, command=self.on_submit)
        self.submit_button.grid(row=7, columnspan=2, pady=(10, 10))

        # Text area to hold the domain names
        self.text_area = Text(self.frame, width=30)
        self.text_area.insert(END, '           Domains\n------------------------------')
        self.text_area.grid(row=8, columnspan=2, pady=(10, 10))

        self.frame.pack()

    def artist_roster_file_handler(self):
        file_name = tkFileDialog.askopenfilename()

        self.artist_roster_filename_text_field.config(text=file_name)

    def old_domain_file_handler(self):
        file_name = tkFileDialog.askopenfilename()

        self.old_domain_filename_text_field.config(text=file_name)

    def open_directory_handler(self):
        directory = tkFileDialog.askdirectory()

        self.new_domain_directory_text_field.config(text=directory)

    def on_submit(self):
        # print("The artist roster filename is: {}".format(self.artist_roster_filename_text_field['text']))
        # print("The old domain filename is: {}".format(self.old_domain_filename_text_field['text']))
        # print("The new domain directory is: {}".format(self.new_domain_directory_text_field['text']))
        # print("The new domain filename is: {}".format(self.new_domain_filename_entry.get()))

        artist_roster_filename_string_var = self.artist_roster_filename_text_field['text']
        old_domain_filename_string_var = self.old_domain_filename_text_field['text']
        new_domain_directory_string_var = self.new_domain_directory_text_field['text']
        new_domain_filename_string_var = self.new_domain_filename_entry.get()

        process_input_and_output(artist_roster_filename_string_var, old_domain_filename_string_var,
                                 new_domain_directory_string_var, new_domain_filename_string_var, self.text_area)

        self.newWindow = Toplevel(self.master)
        self.app = processing_window(self.newWindow)

class processing_window:
    def __init__(self, master):
        self.master = master
        self.frame = Frame(self.master)
        self.processing_label = Label(self.frame, text='Whois lookups are complete!')
        self.processing_label.pack()
        self.quitButton = Button (self.frame, text='Exit', width=25, command=self.exit_program)
        self.quitButton.pack()
        self.frame.pack()

    def exit_program(self):
        root.destroy()

def main():
    global root
    root = Tk()
    app = input_window(root)
    root.mainloop()

def strip_all_text_but_tld(cell_text):

    domain = get_tld(cell_text, fail_silently=True)

    if domain is None:
        return cell_text
    else:
        return domain

def check_if_atl_owned(w):
    atl_owned = '---'

    if w['registrar'] > 0:
        if 'CSC' in w['registrar']:
            atl_owned = 'CSC'
        elif 'Mark' in w['registrar']:
            atl_owned = 'MARK'

    elif w['emails'] > 0:
        if '@atlanticrecords.com' in w['emails']:
            atl_owned = w['registrar']

    elif w['org'] > 0:
        if 'Warner Music' in w['org']:
            atl_owned = w['registrar']

    return atl_owned

def get_exp_date(w):

    if isinstance(w['expiration_date'], list):
        exp_date = w['expiration_date'][0]
    else:
        exp_date = w['expiration_date']
    return exp_date


def do_notes_and_registrar(w):
    info = 'Registrant: '

    if w['org'] > 0:
        if 'Warner Music' in w['org']:
            info += 'Warner Music Group'
        elif 'Atlantic Record' in w['org']:
            info += 'Atlantic Records'
        else:
            info += w['org']

    elif w['name'] > 0:
        info += w['name']

    return info

def create_write_to_workbook():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'ACTIVE_NEW'

    ws1['A1'] = 'Artist'
    ws1['B1'] = 'Digital Marketer'
    ws1['C1'] = 'Domain'
    ws1['D1'] = 'ATL Own?'
    ws1['E1'] = 'Expiration'
    ws1['F1'] = 'Notes/Registrar'

    return wb

def add_whois_info(wb, cell_row_index, info):
    cell_col_index = 3 # starts at Excel cell col 'C'

    ws_name = wb.sheetnames[0]
    ws1 = wb.get_sheet_by_name(ws_name)

    atl_owned_fill = PatternFill(fill_type='solid', fgColor='CCFFCC')
    non_atl_owned_fill = PatternFill(fill_type='solid', fgColor='FF8080')

    for index, i in enumerate(info):
        cell = ws1.cell(column=cell_col_index, row=cell_row_index, value='{}'.format(i))
        if index == 1: # part of the 'info' array pertaining to 'ATL Own?'
            if 'CSC' in i or 'MARK' in i:
                cell.fill = atl_owned_fill
            elif 'Atlantic Record' in info[3] or 'Warner Music' in info[3]: # check registrant/registrar info
                cell.fill = atl_owned_fill
            else:
                cell.fill = non_atl_owned_fill
        cell_col_index += 1

    return wb

def add_non_whois_info(old_sheet, wb):
    ws_name = wb.sheetnames[0]
    new_sheet = wb.get_sheet_by_name(ws_name)

    for row in old_sheet.iter_rows('A{}:B{}'.format(old_sheet.min_row, old_sheet.max_row)):
        for cell in row:
            new_sheet['{}{}'.format(cell.column, cell.row)] = cell.value
    return wb

def verify_artists_and_digital_marketers(wb, artist_roster_filename_string_var):
    artist_roster = load_workbook(filename = artist_roster_filename_string_var)
    roster_sheet_name = artist_roster.sheetnames[0]
    roster_sheet = artist_roster.get_sheet_by_name(roster_sheet_name)

    new_domain_list_sheet_name = wb.sheetnames[0]
    new_domain_list_sheet = wb.get_sheet_by_name(new_domain_list_sheet_name)

    # First, store all of the artists and dms
    artist_to_dm_dict = {} # data structure to keep track of 'digital roster' artists -> dm
    for artist_cell in roster_sheet['C']:
        artist_to_dm_dict[artist_cell.value] = roster_sheet['B{}'.format(artist_cell.row)].value

    domain_artist_to_dm_dict = {} # data structure to keep track of 'new domain book' artists -> dm
    for artist_cell in new_domain_list_sheet['A']:
        domain_artist_to_dm_dict[artist_cell.value] = new_domain_list_sheet['B{}'.format(artist_cell.row)].value

    # Second, loop through and see if artists in 'artist_to_dm_dict' are also contained in 'domain_artist_to_dm_dict'
    for artist, artist_dm in artist_to_dm_dict.iteritems():
        if domain_artist_to_dm_dict.get(artist) == None:
            # Add artist and dm to new workbook sheet
            max_row = new_domain_list_sheet.max_row
            new_domain_list_sheet.cell(row=max_row + 1, column=1, value='{}'.format(artist))
            new_domain_list_sheet.cell(row=max_row + 1, column=2, value='{}'.format(artist_dm))

        elif domain_artist_to_dm_dict.get(artist) != artist_dm:
            # Update new workbook sheet with dm
            for artist_cell in new_domain_list_sheet['A']:
                if artist_cell.value == artist:
                    new_domain_list_sheet.cell(row=artist_cell.row, column=2, value='{}'.format(artist_dm))

    return wb

def add_sort_conditions(wb, cell_row_end):
    ws_name = wb.sheetnames[0]
    ws1 = wb.get_sheet_by_name(ws_name)

    ws1.auto_filter.ref = "A1:F{}".format(cell_row_end)
    ws1.auto_filter.add_sort_condition("E2:E{}".format(cell_row_end))

    return wb

def process_input_and_output(artist_roster_filename_string_var, old_domain_filename_string_var,
                             new_domain_directory_string_var, new_domain_filename_string_var, text_area):
    loaded_wb = load_workbook(filename = old_domain_filename_string_var)
    ws_name = loaded_wb.sheetnames[0]
    ws1 = loaded_wb.get_sheet_by_name(ws_name)
    cell_row_index = 2 # starts at Excel cell row '2'

    write_to_wb = create_write_to_workbook()

    for cell in ws1['C']:
        if cell.value == "Domain":
            continue
        tld = strip_all_text_but_tld(cell.value)

        text_area.insert(END, '{}\n'.format(tld))

        w = whois.whois(tld)

        # print(w)

        if w['domain_name'] is not None:
            # domain = get_domain(w)
            org = check_if_atl_owned(w)
            expiration_date = get_exp_date(w)
            notes = do_notes_and_registrar(w)

        else:
            org = 'MANUAL LOOKUP'
            expiration_date = 'MANUAL LOOKUP'
            notes = 'MANUAL LOOKUP'

        info = []  # structure to hold whois info
        info.extend((tld, org, expiration_date, notes))
        write_to_wb = add_whois_info(write_to_wb, cell_row_index, info)

        cell_row_index += 1

    write_to_wb = add_non_whois_info(ws1, write_to_wb)
    write_to_wb = verify_artists_and_digital_marketers(write_to_wb, artist_roster_filename_string_var)
    write_to_wb = add_sort_conditions(write_to_wb, cell_row_index)
    write_to_wb.save(os.path.join(new_domain_directory_string_var, new_domain_filename_string_var + '.xlsx'))


if __name__ == '__main__':
    main()
